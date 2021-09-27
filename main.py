from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from time import sleep


def get_stock_data(companies,ws):
    
    rows = 2

    for ticker in companies:
    
        url = f'https://finance.yahoo.com/quote/{ticker}'
        html_text = requests.get(url).text

        soup = BeautifulSoup(html_text, "lxml")

        company_name = soup.find('h1', class_ = 'D(ib) Fz(18px)').text
        current_price = soup.find('div', class_ = 'D(ib) Mend(20px)').find_all('span')[0].text
        volume = soup.find('table', class_ = 'W(100%)').find('tbody').find_all('tr')[6].find_all('td')[1].text
        average_volume = soup.find('table', class_ = 'W(100%)').find('tbody').find_all('tr')[7].find_all('td')[1].text
        market_cap = soup.find('table', class_ = 'W(100%) M(0) Bdcl(c)').find('tbody').find_all('tr')[0].find_all('td')[1].text
        beta = soup.find('table', class_ = 'W(100%) M(0) Bdcl(c)').find('tbody').find_all('tr')[1].find_all('td')[1].text
        pe_ratio = soup.find('table', class_ = 'W(100%) M(0) Bdcl(c)').find('tbody').find_all('tr')[2].find_all('td')[1].text

        data = [company_name, current_price, volume, average_volume, market_cap, beta, pe_ratio]

        for column in range(2,ws.max_column + 1):
            ws[ get_column_letter(column) + f'{rows}'] = data[column - 2]

        rows += 1


def save_updated_spreadsheet(wb, ws):
    
    stocks = []

    for row in range(2, ws.max_row + 1):
        stocks.append(ws['A' + f'{row}'].value.upper())

    get_stock_data(stocks,ws)
    wb.save('StocksTemplate.xlsx')
    print ("Updated Spreadsheet successfully saved!")


def main():
    wb = load_workbook(filename = 'StocksTemplate.xlsx')
    ws = wb.active

    print ("How often would you like to update the spreadsheet? Please enter time in minutes:")
    minutes_wait = float(input())

    while True:
        print("Program is running")
        save_updated_spreadsheet(wb, ws)
        print (f'Program will run again in {minutes_wait} minutes')
        sleep(minutes_wait * 60)
    

main()